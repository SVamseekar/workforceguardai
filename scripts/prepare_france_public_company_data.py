#!/usr/bin/env python3
"""Extract France employer-level public company benchmark data from the official Egapro XLSX."""
from __future__ import annotations

import argparse
import json
import os
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, Optional

import pandas as pd
import requests
from openpyxl import load_workbook


SOURCE_URL = "https://www.data.gouv.fr/api/1/datasets/r/d434859f-8d3b-4381-bcdb-ec9200653ae6"
SOURCE_PUBLISHED_AT = "2026-04-13"
DISCLOSURE_FRAMEWORK = "Index Egalite Professionnelle F/H (France)"
PROVENANCE_TEMPLATE = (
    "Official data.gouv.fr XLSX resource d434859f-8d3b-4381-bcdb-ec9200653ae6. "
    "UES-level declarations where applicable."
)

BENCHMARK_TERMS = {
    "totalenergies": ["TotalEnergies", "TOTALENERGIES"],
    "edf": ["EDF", "Electricite de France", "ELECTRICITE DE FRANCE"],
}

BENCHMARK_NAME_PREFERENCES = {
    "totalenergies": [
        "TOTALENERGIES SE",
        "UES TOTALENERGIES",
        "TOTALENERGIES",
    ],
    "edf": [
        "EDF SA",
        "ELECTRICITE DE FRANCE SA",
        "ELECTRICITE DE FRANCE",
        "EDF",
    ],
}

MISSING_VALUE_MARKERS = {"", "-", "NC", "N.C.", "N/A", "NA", "NAN", "NONE", "NULL", "VIDE"}


def remove_accents(value: Any) -> str:
    text = str(value) if value is not None else ""
    return "".join(
        character
        for character in unicodedata.normalize("NFD", text)
        if unicodedata.category(character) != "Mn"
    )


def normalize_header(value: Any) -> str:
    normalized = remove_accents(value).strip().lower()
    normalized = re.sub(r"[^a-z0-9_ ]", "", normalized)
    normalized = re.sub(r"\s+", "_", normalized)
    return normalized.strip("_")


def normalize_company_name(value: Any) -> str:
    normalized = remove_accents(value).strip().upper()
    normalized = re.sub(r"[^A-Z0-9 ]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized.strip()


def is_missing_marker(value: Any) -> bool:
    if value is None or pd.isna(value):
        return True
    return str(value).strip().upper() in MISSING_VALUE_MARKERS


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def download_file(url: str, destination: Path, timeout_seconds: int = 120) -> None:
    if destination.exists():
        return

    ensure_parent_dir(destination)
    response = requests.get(url, stream=True, timeout=timeout_seconds)
    response.raise_for_status()

    with destination.open("wb") as handle:
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                handle.write(chunk)


def detect_header_row(preview_frame: pd.DataFrame) -> int:
    best_row_index = 0
    best_score = float("-inf")
    keywords = (
        "siren",
        "raison",
        "sociale",
        "entreprise",
        "index",
        "indicateur",
        "remuneration",
        "maternite",
    )

    for row_index in range(min(len(preview_frame), 15)):
        values = [normalize_header(value) for value in preview_frame.iloc[row_index].tolist()]
        non_empty_count = sum(1 for value in values if value)
        keyword_hits = sum(any(keyword in value for keyword in keywords) for value in values)
        unique_count = len({value for value in values if value})
        score = keyword_hits * 100 + non_empty_count * 5 + unique_count

        if score > best_score:
            best_score = score
            best_row_index = row_index

    return best_row_index


def inspect_workbook(workbook_path: Path) -> tuple[list[str], dict[str, dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    metadata: dict[str, dict[str, Any]] = {}

    for sheet_name in sheet_names:
        preview = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None, nrows=30)
        header_row = detect_header_row(preview)
        sheet_frame = pd.read_excel(workbook_path, sheet_name=sheet_name, header=header_row)
        metadata[sheet_name] = {
            "header_row": header_row,
            "columns": [str(column) for column in sheet_frame.columns.tolist()],
            "num_rows": int(len(sheet_frame)),
            "num_cols": int(len(sheet_frame.columns)),
        }

    return sheet_names, metadata


def detect_main_sheet(sheet_names: Iterable[str], metadata: dict[str, dict[str, Any]]) -> str:
    best_sheet = next(iter(sheet_names))
    best_score = float("-inf")

    for sheet_name in sheet_names:
        info = metadata[sheet_name]
        normalized_columns = [normalize_header(column) for column in info["columns"]]
        score = info["num_rows"] * 5 + info["num_cols"] * 2
        if any("siren" in column for column in normalized_columns):
            score += 2000
        if any("raison" in column or "entreprise" in column for column in normalized_columns):
            score += 1000
        if any("index" in column or "indicateur" in column for column in normalized_columns):
            score += 500

        if score > best_score:
            best_score = score
            best_sheet = sheet_name

    return best_sheet


def _best_candidate(
    candidates: list[str],
    normalized_columns: dict[str, str],
    keywords: list[str],
    *,
    exclude_keywords: Optional[list[str]] = None,
) -> Optional[str]:
    scored: list[tuple[int, str]] = []
    exclusions = exclude_keywords or []

    for original_column in candidates:
        normalized = normalized_columns[original_column]
        if any(exclusion in normalized for exclusion in exclusions):
            continue

        score = 0
        for index, keyword in enumerate(keywords):
            if normalized == keyword:
                score += 1000 - index
            elif normalized.startswith(keyword):
                score += 800 - index
            elif keyword in normalized:
                score += 500 - index

        if score > 0:
            scored.append((score, original_column))

    if not scored:
        return None

    scored.sort(key=lambda item: (-item[0], len(item[1]), item[1]))
    return scored[0][1]


def detect_column_mapping(frame: pd.DataFrame) -> tuple[dict[str, Optional[str]], list[str]]:
    normalized_columns = {column: normalize_header(column) for column in frame.columns}
    column_names = list(frame.columns)

    mapping = {
        "employer_name": _best_candidate(
            column_names,
            normalized_columns,
            [
                "raison_sociale",
                "nom_entreprise",
                "entreprise",
                "raison",
                "societe",
            ],
        ),
        "employer_id": _best_candidate(
            column_names,
            normalized_columns,
            [
                "siren",
                "sirene",
            ],
        ),
        "reporting_year": _best_candidate(
            column_names,
            normalized_columns,
            [
                "annee",
                "annee_index",
                "annee_declaration",
                "publication",
                "exercice",
            ],
        ),
        "overall_score": _best_candidate(
            column_names,
            normalized_columns,
            [
                "index",
                "note_index",
                "score_global",
                "resultat_index",
            ],
            exclude_keywords=["indicateur", "ecart"],
        ),
        "pay_gap_indicator_score": _best_candidate(
            column_names,
            normalized_columns,
            [
                "indicateur_1_ecart_remuneration",
                "indicateur_1_ecart_de_remuneration",
                "ecart_remuneration_points",
                "points_ecart_remuneration",
                "ecart_de_remuneration",
                "ecart_remuneration",
            ],
        ),
        "pay_gap_percent": _best_candidate(
            column_names,
            normalized_columns,
            [
                "pourcentage_ecart_remuneration",
                "pourcentage_ecart_de_remuneration",
                "ecart_remuneration_pourcentage",
                "pourcentage_ecart_de_remuneration",
                "ecart_remuneration_percent",
            ],
        ),
        "promotion_gap_score": _best_candidate(
            column_names,
            normalized_columns,
            [
                "indicateur_3_ecart_promotions",
                "indicateur_3_promotions",
                "note_ecart_taux_de_promotion",
                "points_ecart_promotions",
                "ecart_promotions",
                "promotions",
            ],
        ),
        "maternity_return_score": _best_candidate(
            column_names,
            normalized_columns,
            [
                "indicateur_4_retour_conge_maternite",
                "indicateur_4_maternite",
                "augmentation_apres_maternite",
                "maternite",
            ],
        ),
        "top10_underrepresented_gender_score": _best_candidate(
            column_names,
            normalized_columns,
            [
                "indicateur_5_top_10",
                "indicateur_5_top10",
                "femmes_parmi_les_10_plus_hautes_remunerations",
                "note_hautes_remunerations",
                "points_top_10",
                "top10",
            ],
        ),
        "effectif": _best_candidate(
            column_names,
            normalized_columns,
            [
                "tranche_effectifs",
                "tranche_effectif",
                "effectifs",
                "effectif",
                "nombre_salaries",
            ],
        ),
    }

    matched_columns = {column for column in mapping.values() if column}
    unmatched_columns = [str(column) for column in frame.columns if column not in matched_columns]
    return mapping, unmatched_columns


def _row_signature(row: pd.Series, columns: list[str]) -> tuple[str, ...]:
    return tuple("" if pd.isna(row[column]) else str(row[column]) for column in columns)


def find_company_matches(frame: pd.DataFrame, mapping: dict[str, Optional[str]]) -> pd.DataFrame:
    name_column = mapping["employer_name"] or str(frame.columns[0])
    working_frame = frame.copy()
    working_frame["_normalized_name"] = working_frame[name_column].map(normalize_company_name)

    grouped_matches: dict[tuple[tuple[str, ...], str], dict[str, Any]] = {}
    source_columns = list(frame.columns)

    for benchmark_key, terms in BENCHMARK_TERMS.items():
        normalized_terms = [normalize_company_name(term) for term in terms]
        mask = working_frame["_normalized_name"].map(
            lambda value: any(term in value for term in normalized_terms)
        )

        for _, row in working_frame[mask].iterrows():
            signature = _row_signature(row, source_columns)
            group_key = (signature, benchmark_key)
            record = grouped_matches.get(group_key)

            if record is None:
                record = {column: row[column] for column in source_columns}
                record["benchmark_key"] = benchmark_key
                record["matched_search_terms"] = []
                grouped_matches[group_key] = record

            for term in terms:
                if term not in record["matched_search_terms"]:
                    record["matched_search_terms"].append(term)

    records = []
    for record in grouped_matches.values():
        record["matched_search_terms"] = "|".join(sorted(record["matched_search_terms"]))
        records.append(record)

    if not records:
        return pd.DataFrame(columns=list(frame.columns) + ["benchmark_key", "matched_search_terms"])

    return pd.DataFrame(records).reset_index(drop=True)


def parse_numeric(value: Any) -> Optional[float]:
    if is_missing_marker(value):
        return None

    text = str(value).strip()
    text = text.replace("\xa0", " ")
    text = text.replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None

    try:
        return float(match.group(0))
    except ValueError:
        return None


def parse_year(value: Any) -> Optional[int]:
    if is_missing_marker(value):
        return None

    match = re.search(r"(20\d{2}|19\d{2})", str(value))
    if not match:
        return None
    return int(match.group(1))


def parse_effectif_upper_bound(value: Any) -> float:
    if is_missing_marker(value):
        return 0.0

    text = remove_accents(str(value)).upper()
    if "PLUS" in text or "+" in text:
        digits = [int(number) for number in re.findall(r"\d+", text)]
        return float(max(digits)) if digits else 1_000_000.0

    digits = [int(number) for number in re.findall(r"\d+", text)]
    if not digits:
        return 0.0
    return float(max(digits))


def benchmark_name_preference_score(name: Any, benchmark_key: str) -> float:
    normalized_name = normalize_company_name(name)
    preferences = BENCHMARK_NAME_PREFERENCES[benchmark_key]
    score = 0.0

    for index, preferred_name in enumerate(preferences):
        normalized_preference = normalize_company_name(preferred_name)
        if normalized_name == normalized_preference:
            score += 400 - index * 10
        elif normalized_name.startswith(normalized_preference):
            score += 250 - index * 10
        elif normalized_preference in normalized_name:
            score += 120 - index * 10

    if benchmark_key == "totalenergies":
        for penalty_marker in ("MARKETING", "SERVICES", "HOLDING", "PETROCHIMIE", "RAFFINAGE"):
            if penalty_marker in normalized_name and "TOTALENERGIES SE" not in normalized_name:
                score -= 25
    if benchmark_key == "edf":
        for penalty_marker in ("RENOUVELABLES", "TRADING", "SOLUTIONS", "DALKIA"):
            if penalty_marker in normalized_name and "EDF SA" not in normalized_name:
                score -= 25

    return score


def completeness_score(row: pd.Series, mapping: dict[str, Optional[str]]) -> int:
    fields = [
        "overall_score",
        "pay_gap_indicator_score",
        "promotion_gap_score",
        "maternity_return_score",
        "top10_underrepresented_gender_score",
    ]
    score = 0
    for field in fields:
        column = mapping.get(field)
        if column and not is_missing_marker(row.get(column)):
            score += 1
    return score


def score_candidate_row(row: pd.Series, mapping: dict[str, Optional[str]]) -> float:
    score = 0.0

    benchmark_key = row.get("benchmark_key", "")
    employer_name = row.get(mapping.get("employer_name", ""))
    score += benchmark_name_preference_score(employer_name, benchmark_key)

    reporting_year = parse_year(row.get(mapping.get("reporting_year", "")))
    if reporting_year is not None:
        score += reporting_year * 100

    score += completeness_score(row, mapping) * 60

    effectif_value = parse_effectif_upper_bound(row.get(mapping.get("effectif", "")))
    score += effectif_value / 1000

    overall_score = parse_numeric(row.get(mapping.get("overall_score", "")))
    if overall_score is not None:
        score += overall_score * 3

    return score


def build_selection_reason(row: pd.Series, mapping: dict[str, Optional[str]]) -> str:
    reporting_year = parse_year(row.get(mapping.get("reporting_year", "")))
    effectif_value = parse_effectif_upper_bound(row.get(mapping.get("effectif", "")))
    overall_score = parse_numeric(row.get(mapping.get("overall_score", "")))
    completeness = completeness_score(row, mapping)
    name_score = benchmark_name_preference_score(
        row.get(mapping.get("employer_name", "")),
        row.get("benchmark_key", ""),
    )

    parts = [f"name preference {name_score:.1f}", f"indicator completeness {completeness}"]
    if reporting_year is not None:
        parts.append(f"reporting year {reporting_year}")
    if effectif_value:
        parts.append(f"size upper bound {int(effectif_value)}")
    if overall_score is not None:
        parts.append(f"overall score {overall_score:g}")

    return "; ".join(parts)


def select_best_rows(matches: pd.DataFrame, mapping: dict[str, Optional[str]]) -> pd.DataFrame:
    if matches.empty:
        return pd.DataFrame(columns=list(matches.columns) + ["selection_score", "selected_for_benchmark", "selection_reason"])

    scored = matches.copy()
    scored["selection_score"] = scored.apply(lambda row: score_candidate_row(row, mapping), axis=1)
    scored["selection_reason"] = scored.apply(lambda row: build_selection_reason(row, mapping), axis=1)
    scored["selected_for_benchmark"] = False

    selected_rows = []
    for benchmark_key in BENCHMARK_TERMS:
        group = scored[scored["benchmark_key"] == benchmark_key]
        if group.empty:
            continue

        best_index = group["selection_score"].idxmax()
        best_row = scored.loc[best_index].copy()
        best_row["selected_for_benchmark"] = True
        selected_rows.append(best_row)

    if not selected_rows:
        return scored.iloc[0:0].copy()

    return pd.DataFrame(selected_rows).reset_index(drop=True)


def build_normalized_output(
    selected_rows: pd.DataFrame,
    mapping: dict[str, Optional[str]],
    *,
    source_url: str,
    source_published_at: str,
    source_retrieved_at: str,
) -> pd.DataFrame:
    if selected_rows.empty:
        return pd.DataFrame(
            columns=[
                "country_code",
                "employer_name",
                "employer_id",
                "reporting_year",
                "disclosure_framework",
                "overall_score",
                "pay_gap_indicator_score",
                "pay_gap_percent",
                "promotion_gap_score",
                "maternity_return_score",
                "top10_underrepresented_gender_score",
                "source_url",
                "source_published_at",
                "source_retrieved_at",
                "provenance_notes",
            ]
        )

    normalized = pd.DataFrame()
    normalized.index = selected_rows.index
    normalized["country_code"] = pd.Series("FR", index=selected_rows.index, dtype="string")
    normalized["employer_name"] = (
        selected_rows[mapping["employer_name"]].astype("string")
        if mapping["employer_name"]
        else pd.Series(dtype="string")
    )
    normalized["employer_id"] = (
        selected_rows[mapping["employer_id"]].astype("string")
        if mapping["employer_id"]
        else pd.Series(dtype="string")
    )
    normalized["reporting_year"] = (
        selected_rows[mapping["reporting_year"]].map(parse_year)
        if mapping["reporting_year"]
        else pd.Series(dtype="Int64")
    )
    normalized["disclosure_framework"] = pd.Series(
        DISCLOSURE_FRAMEWORK,
        index=selected_rows.index,
        dtype="string",
    )
    normalized["overall_score"] = (
        selected_rows[mapping["overall_score"]].map(parse_numeric)
        if mapping["overall_score"]
        else pd.Series(dtype="float")
    )
    normalized["pay_gap_indicator_score"] = (
        selected_rows[mapping["pay_gap_indicator_score"]].map(parse_numeric)
        if mapping["pay_gap_indicator_score"]
        else pd.Series(dtype="float")
    )
    normalized["pay_gap_percent"] = (
        selected_rows[mapping["pay_gap_percent"]].map(parse_numeric)
        if mapping["pay_gap_percent"]
        else pd.Series(dtype="float")
    )
    normalized["promotion_gap_score"] = (
        selected_rows[mapping["promotion_gap_score"]].map(parse_numeric)
        if mapping["promotion_gap_score"]
        else pd.Series(dtype="float")
    )
    normalized["maternity_return_score"] = (
        selected_rows[mapping["maternity_return_score"]].map(parse_numeric)
        if mapping["maternity_return_score"]
        else pd.Series(dtype="float")
    )
    normalized["top10_underrepresented_gender_score"] = (
        selected_rows[mapping["top10_underrepresented_gender_score"]].map(parse_numeric)
        if mapping["top10_underrepresented_gender_score"]
        else pd.Series(dtype="float")
    )
    normalized["source_url"] = pd.Series(source_url, index=selected_rows.index, dtype="string")
    normalized["source_published_at"] = pd.Series(source_published_at, index=selected_rows.index, dtype="string")
    normalized["source_retrieved_at"] = pd.Series(source_retrieved_at, index=selected_rows.index, dtype="string")
    normalized["provenance_notes"] = pd.Series(PROVENANCE_TEMPLATE, index=selected_rows.index, dtype="string")

    return normalized


def extract_france_public_company_data(
    *,
    source_url: str,
    raw_file: Path,
    headers_json: Path,
    matches_csv: Path,
    candidates_csv: Path,
    selected_csv: Path,
    normalized_csv: Path,
    audit_json: Path,
    source_published_at: str = SOURCE_PUBLISHED_AT,
    download: bool = True,
) -> dict[str, Any]:
    if download:
        download_file(source_url, raw_file)

    sheet_names, metadata = inspect_workbook(raw_file)
    main_sheet = detect_main_sheet(sheet_names, metadata)
    header_row = metadata[main_sheet]["header_row"]
    frame = pd.read_excel(raw_file, sheet_name=main_sheet, header=header_row)
    mapping, unmatched_columns = detect_column_mapping(frame)
    matches = find_company_matches(frame, mapping)
    selected_rows = select_best_rows(matches, mapping)
    scored_candidates = matches.copy()
    if not scored_candidates.empty:
        scored_candidates["selection_score"] = scored_candidates.apply(
            lambda row: score_candidate_row(row, mapping),
            axis=1,
        )
        scored_candidates["selection_reason"] = scored_candidates.apply(
            lambda row: build_selection_reason(row, mapping),
            axis=1,
        )
        selected_signatures = {
            (_row_signature(row, list(frame.columns)), row["benchmark_key"])
            for _, row in selected_rows.iterrows()
        }
        scored_candidates["selected_for_benchmark"] = scored_candidates.apply(
            lambda row: (_row_signature(row, list(frame.columns)), row["benchmark_key"]) in selected_signatures,
            axis=1,
        )
    else:
        scored_candidates["selection_score"] = pd.Series(dtype="float")
        scored_candidates["selection_reason"] = pd.Series(dtype="string")
        scored_candidates["selected_for_benchmark"] = pd.Series(dtype="boolean")

    source_retrieved_at = datetime.now(timezone.utc).isoformat()
    normalized = build_normalized_output(
        selected_rows,
        mapping,
        source_url=source_url,
        source_published_at=source_published_at,
        source_retrieved_at=source_retrieved_at,
    )

    ensure_parent_dir(headers_json)
    headers_json.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")

    ensure_parent_dir(matches_csv)
    matches.to_csv(matches_csv, index=False, encoding="utf-8")
    ensure_parent_dir(candidates_csv)
    scored_candidates.to_csv(candidates_csv, index=False, encoding="utf-8")
    ensure_parent_dir(selected_csv)
    selected_rows.to_csv(selected_csv, index=False, encoding="utf-8")
    ensure_parent_dir(normalized_csv)
    normalized.to_csv(normalized_csv, index=False, encoding="utf-8")

    audit_payload = {
        "chosen_sheet": main_sheet,
        "detected_header_row": header_row,
        "mapped_columns": {field: column for field, column in mapping.items() if column},
        "unmatched_columns": unmatched_columns,
        "num_candidate_rows_total": int(len(matches)),
        "num_totalenergies_candidates": int((matches.get("benchmark_key") == "totalenergies").sum()) if not matches.empty else 0,
        "num_edf_candidates": int((matches.get("benchmark_key") == "edf").sum()) if not matches.empty else 0,
        "num_selected_benchmarks": int(len(selected_rows)),
        "source_url": source_url,
        "source_published_at": source_published_at,
        "source_retrieved_at": source_retrieved_at,
    }
    ensure_parent_dir(audit_json)
    audit_json.write_text(json.dumps(audit_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "metadata": metadata,
        "chosen_sheet": main_sheet,
        "header_row": header_row,
        "mapping": mapping,
        "unmatched_columns": unmatched_columns,
        "matches": matches,
        "candidates": scored_candidates,
        "selected": selected_rows,
        "normalized": normalized,
        "audit": audit_payload,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Extract France public company benchmark rows from the official Egapro XLSX")
    parser.add_argument("--source-url", default=SOURCE_URL)
    parser.add_argument("--raw-file", type=Path, default=Path("data/public_company_raw/france/france_index_raw.xlsx"))
    parser.add_argument("--headers-json", type=Path, default=Path("data/public_company_meta/france/france_sheet_headers.json"))
    parser.add_argument("--matches-csv", type=Path, default=Path("data/public_company_meta/france/france_company_matches.csv"))
    parser.add_argument("--candidates-csv", type=Path, default=Path("data/public_company_meta/france/france_benchmark_candidates.csv"))
    parser.add_argument("--selected-csv", type=Path, default=Path("data/public_company_meta/france/france_selected_benchmarks.csv"))
    parser.add_argument("--normalized-csv", type=Path, default=Path("data/public_company/france_public_company_benchmark.csv"))
    parser.add_argument("--audit-json", type=Path, default=Path("data/public_company_meta/france/france_extraction_audit.json"))
    parser.add_argument("--source-published-at", default=SOURCE_PUBLISHED_AT)
    parser.add_argument(
        "--skip-download",
        action="store_true",
        help="Use an existing local XLSX instead of downloading from the source URL.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    result = extract_france_public_company_data(
        source_url=args.source_url,
        raw_file=args.raw_file,
        headers_json=args.headers_json,
        matches_csv=args.matches_csv,
        candidates_csv=args.candidates_csv,
        selected_csv=args.selected_csv,
        normalized_csv=args.normalized_csv,
        audit_json=args.audit_json,
        source_published_at=args.source_published_at,
        download=not args.skip_download,
    )

    print("France public company extraction complete.")
    print(f"Chosen sheet: {result['chosen_sheet']}")
    print(f"Header row: {result['header_row']}")
    print(f"Candidate rows: {len(result['matches'])}")
    print(f"Selected benchmark rows: {len(result['selected'])}")
    print(f"Normalized output: {args.normalized_csv}")


if __name__ == "__main__":
    main()
